import re
from collections import namedtuple
from openpyxl.utils.cell import cols_from_range, get_column_letter, column_index_from_string
import excelsheet2html.formula_parser as formula_parser

# openpyxl.utils.cell.cols_from_range(range_string)

__all__ = ['read_formulas']


def read_formulas(glb, ws):
    Formulas = namedtuple(
        "Formulas", "primary_cells secondary_cells formula_cells secondary_cells_1 secondary_cells_2")
    primary_cell = []
    formula_cells = []
    cells = list(cols_from_range(glb.values))
    for col_cells in cells:
        for cell_coord in col_cells:
            if ws[cell_coord].data_type == 'f':
                formula_cells.append(ws[cell_coord])
            if ws[cell_coord].data_type == 'n':
                primary_cell.append(ws[cell_coord])

    # output is parent child struct of cell
    # eg. [('E4', ('+1*E5', '+1*E6', '+1*E7'))]
    output = []
    for cell in formula_cells:
        n = cell.coordinate
        frm = formula_parser.parsing_formula(cell.value)
        output.append((n, frm))

    """
    replace  constants (in =A4+6 six is a constant) with fake cells like XXX1, XXX2...
    Cells and coresponding constant value is added to ws object  
    """
    constants_counter = 1
    for i, (c, f) in enumerate(output):
        for idx, expr in enumerate(f):
            if not any(c.isalpha() for c in expr):
                # print(i, idx, expr, id(f))
                n = float(expr[expr.index('*')+1:])
                ws['XXX' + str(constants_counter)] = n
                primary_cell.append(ws['XXX' + str(constants_counter)])
                output[i][1][idx] = expr[:expr.index(
                    '*')+1] + 'XXX' + str(constants_counter)
                constants_counter += 1

    # flattend list of tupple. It is flattened output
    # eg. [('E4', '+1', 'E5'),('E4', '+1', 'E6')]
    # 1-parent cell, 2-multiplier, 3-child cell
    flattend = []
    wtorne = set()
    for e in output:
        for c in e[1]:
            wtorne.add(e[0])
            flattend.append((e[0], *c.split('*')))

    # flat parent child structure with only primary child cells
    # 1 - parent cell
    # 2 - child cell (only primary)
    # 3 - multipler
    # 4 - iterationa or depth level. It tells at what iteration parent cell should be calculated
    out_all = []

    def _recursive_shit(_for, _what, m, iter=1):
        small_tbl = []
        if _what in wtorne:
            small_tbl = [x for x in flattend if _what == x[0]]
            for e in small_tbl:
                _recursive_shit(_for, e[2], float(e[1]) * m, iter+1)
        else:
            out_all.append((_for, _what, m, iter))

    for x in flattend:
        _recursive_shit(_for=x[0], _what=x[2], m=float(x[1]))

    p_cells = dict()
    for x in primary_cell:
        p_cells[x.coordinate] = x.value
    f_cells = dict()
    for x in formula_cells:
        f_cells[x.coordinate] = x.value

    f_cells_ory = dict()
    for x in formula_cells:
        f_cells_ory[x.coordinate] = 0

    s1 = [{"parent": x[0], "child": x[2],
           "multiplier": float(x[1])} for x in flattend]

    s2 = [{"parent": x[0], "child": x[1],
           "multiplier": float(x[2]), "iteration": int(x[3])} for x in out_all]
    glb.formula_cells = f_cells_ory
    result = Formulas(
        p_cells,
        f_cells,
        f_cells_ory,
        s1,
        s2
    )
    return result

import re
from openpyxl.utils.cell import column_index_from_string, get_column_letter
from excelsheet2html.config.constants import *
from collections import namedtuple


__all__ = ['set_sheet_properties']


def _get_named_range_adress(wb: object, name: str) -> str:
    res = list(wb.defined_names[name].destinations)[0][1].replace('$', '')
    if len(res.split(':')) == 1:
        res = res + ":" + res
    return res


def _get_surrouding_areas(labels, headers, values, area):
    """
    return namedtuple containing areas that souround table: top - above header
    left - on left of the table
    right - on the right of the table
    on bottome - below table
    ondead_area - is area above label and on left of header
    """
    Surrouding_areas = namedtuple(
        "Surrouding_areas", "ontop onbottom onleft onright ondead_area")
    left_top_letter = re.findall('[A-Z]+', labels.split(':')[0])[0]
    bottom_right_letter = re.findall('[A-Z]+', headers.split(':')[1])[0]
    bottom_right_number = int(re.findall('\d+', headers.split(':')[1])[0]) - 1
    ontop = f"{left_top_letter}1:{bottom_right_letter}{bottom_right_number}"

    left_top_letter = re.findall('[A-Z]+', labels.split(':')[0])[0]
    left_top_number = int(re.findall('\d+', labels.split(':')[1])[0]) + 1
    bottom_right_letter = re.findall('[A-Z]+', values.split(':')[1])[0]
    bottom_right_number = int(re.findall('\d+', area.split(':')[1])[0])
    onbottom = f"{left_top_letter}{left_top_number}:{bottom_right_letter}{bottom_right_number}"

    right_bottom_letter = get_column_letter(column_index_from_string(
        re.findall('[A-Z]+', labels.split(':')[0])[0])-1)
    right_bottom_number = int(re.findall('\d+', area.split(':')[1])[0])
    onleft = f"A1:{right_bottom_letter}{right_bottom_number}"

    left_top_letter = get_column_letter(column_index_from_string(
        re.findall('[A-Z]+', headers.split(':')[1])[0])+1)
    right_bottom_letter = re.findall('[A-Z]+', area.split(':')[1])[0]
    right_bottom_number = int(re.findall('\d+', area.split(':')[1])[0])
    onright = f"{left_top_letter}1:{right_bottom_letter}{right_bottom_number}"

    left_top_letter = re.findall('[A-Z]+', labels.split(':')[0])[0]
    left_top_number = int(re.findall('\d+', headers.split(':')[0])[0])
    right_bottom_letter = re.findall('[A-Z]+', labels.split(':')[1])[0]
    right_bottom_number = int(re.findall('\d+', headers.split(':')[1])[0])
    ondead_area = f"{left_top_letter}{left_top_number}:{right_bottom_letter}{right_bottom_number}"

    return Surrouding_areas(ontop, onbottom, onleft, onright, ondead_area)


def set_sheet_properties(glb, wb: 'object', sheetname: 'str' = None) -> 'dict ':
    """
    function: dictionary with properties such like:
    area, column_width, list of objects of merged cells
    """

    ws = wb[sheetname or wb.sheetnames[0]]

    area = _get_named_range_adress(wb, 'area')
    labels = _get_named_range_adress(wb, 'labels')
    headers = _get_named_range_adress(wb, 'headers')
    values = _get_named_range_adress(wb, 'values')

    glb.area = area
    glb.labels = labels
    glb.headers = headers
    glb.values = values

    surroudin_areas = _get_surrouding_areas(labels, headers, values, area)
    glb.ontop = surroudin_areas.ontop
    glb.onbottom = surroudin_areas.onbottom
    glb.onleft = surroudin_areas.onleft
    glb.onright = surroudin_areas.onright
    glb.ondead_area = surroudin_areas.ondead_area

    num_of_rows = int(re.findall('\d+', area.split(':')[1])[0])
    num_of_columns = column_index_from_string(
        re.findall('[A-Z]+', area.split(':')[1])[0])
    last_row = int(re.findall('\d+', labels.split(':')[1])[0])
    first_row = int(re.findall('\d+', labels.split(':')[0])[0])
    right_column = column_index_from_string(
        re.findall('[A-Z]+', values.split(':')[1])[0])

    merged_cells = dict()
    for val in ws.merged_cells.ranges:
        colspan = val.bounds[2] - val.bounds[0] + 1
        rowspan = val.bounds[3] - val.bounds[1] + 1
        merged_cells[val.start_cell.coordinate] = {}
        inner_dict = merged_cells[val.start_cell.coordinate]
        if colspan > 1:
            inner_dict['colspan'] = colspan
        if rowspan > 1:
            inner_dict['rowspan'] = rowspan

    glb.num_of_rows = num_of_rows
    glb.num_of_columns = num_of_columns
    glb.last_row = last_row
    glb.first_row = first_row
    glb.right_column = right_column
    glb.merged_cells = merged_cells
    glb.grid_column_width = INDEX_COLUMN_WIDTH
    glb.column_width = {}

    # column_dimensions does't return all columns, in this loop we check i column is retured. And if not with of previous one is taken
    col_width = round(96 * 0.89)
    for i in range(1, num_of_columns + 1):
        if ws.column_dimensions.get(get_column_letter(i), None):
            col = ws.column_dimensions.get(get_column_letter(i))
            if col.customWidth:
                width = round(col.width / 10.0, 2)
            col_width = round(96 * width)
            glb.column_width.update({get_column_letter(i): col_width})
        else:
            glb.column_width.update({get_column_letter(i): col_width})
    glb.total_width = sum(
        [x for x in glb.column_width.values()]) + INDEX_COLUMN_WIDTH

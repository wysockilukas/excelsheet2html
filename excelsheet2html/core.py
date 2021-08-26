"""Read excel file and convert it to html string"""

import openpyxl
import excelsheet2html.helper_functions as my_func
from excelsheet2html.stylesheet.stylesheet import StyleSheet
from excelsheet2html.dom.dom import DOM
from excelsheet2html.config.config import MyGlobals
import excelsheet2html.config.config as config
from excelsheet2html.js_script.script import get_js_script
from collections import namedtuple


def _render_data_to_html(*, css, body, script):
    """
    return string with complete html page
    """
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Title</title>
        {css}
    </head>
    <body>
        {body}
        {script}
    </body>
    </html>
    """
    return html


def excelsheet2html(filepath, sheetname=None, class_name_prefix="c_", read_borders=False, number_format=None):
    """
    Read excel file and convert table from excel sheet into  tupple of strings with: whole html , style part, js part and body part

    :param filepath: The path to open or a file-like object
    :type filename: string or a file-like object

    :param sheetname: Sheet name, if not provided first is selected
    :type sheetname: string

    :param class_name_prefix: Prefix for the class name in the css section. The default value is c_, so in the <style> section there are names such as c_label
    :type class_name_prefix: string   

    :param read_borders: Defaults to false. False means that the cell border properties are not taken from Excel, but the default style is used. If true, the properties of borders from excel are taken
    :type read_borders: boolean 

    :param number_format: Defalut is None. Can be None, space or commma. It tells how numbers should be formated on html page
    :type number_format: string       

    rtype: :namedtuple: `Output`  

    .. note::

        The table in the excel sheet must contain named areas and these names must be specifically:
        area - the area of the sheet to be converted to html
        headers - table headers
        labels - table row labels
        values - an area with numeric values and formulas

    """
    Output = namedtuple("Output", "style table script html")
    """ Final returned strings.
    .. py:attribute:: style
        The style attribute contain <style> tag content
    .. py:attribute:: table
        The table attribute has <table> content
    .. py:attribute:: script
        The script attribute has <script> js conteny
    .. py:attribute:: html
        The html attribute is whole html page            
    """
    wb = openpyxl.load_workbook(filepath)
    wb_data = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheetname or wb.sheetnames[0]]
    ws_data = wb_data[sheetname or wb_data.sheetnames[0]]

    # declare object glb - for global settings, it will be pass down to other functions and classes. Some of them will modify the objecy
    glb = MyGlobals()

    # set_sheet_properties, it will add initial properties to glb object
    my_func.set_sheet_properties(glb, wb, sheetname)
    glb.theme_colors = config.set_theme_colors(wb)

    # return named tuple with: primary_cells, secondary_cells, formula_cells, secondary_cells_1, secondary_cells_2
    formulas = my_func.read_formulas(glb, ws)

    if read_borders:
        glb.read_borders = read_borders

    stylesheet = StyleSheet(glb, class_name_prefix)

    dom = DOM(glb, ws_data, stylesheet)
    table_html = dom.render_table()
    stylesheet_css = stylesheet.get_global_css()
    script = get_js_script(class_name_prefix, formulas, number_format, 0)
    html = _render_data_to_html(
        css=stylesheet_css, body=table_html, script=script)

    output = Output(
        stylesheet_css,
        table_html,
        script,
        html
    )
    return output

    # print(stylesheet.get_global_css())
